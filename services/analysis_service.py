import pandas as pd
import matplotlib.pyplot as plt
from math import pi
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
import matplotlib.font_manager as fm
import io
import os
import tempfile
import re

# 폰트 경로 설정 (qr_service와 동일하게 맞춤)
FONT_PATH = "static/fonts/malgunbd.ttf"


def sanitize_sheet_title(title):
    if isinstance(title, bytes):
        title = title.decode('utf-8')
    invalid_chars = re.compile(r'[\\/*?\[\]:]')
    sanitized_title = invalid_chars.sub('', str(title))
    return sanitized_title[:31]


# 폰트 불러오기 함수 (QR 코드 로직과 유사하게 변경)
def load_custom_font():
    """지정된 경로의 폰트를 로드하여 FontProperties 객체를 반환합니다."""
    try:
        # 1. 지정된 경로에 폰트가 있는지 확인
        if os.path.exists(FONT_PATH):
            return fm.FontProperties(fname=FONT_PATH)

        # 2. 없으면 시스템 폰트 중 '맑은 고딕' 시도 (윈도우 개발 환경용)
        # 리눅스라면 나눔고딕 등이 설치되어 있어야 함
        return fm.FontProperties(family='Malgun Gothic')
    except:
        # 3. 최후의 수단 (기본 폰트, 한글 깨질 수 있음)
        print("한글 폰트 로드 실패. 기본 폰트를 사용합니다.")
        return fm.FontProperties()


# [기능 1] 평균 계산 로직 (변경 없음)
def calculate_trimmed_mean_logic(df: pd.DataFrame) -> io.BytesIO:
    if len(df.columns) < 3:
        raise ValueError("데이터 컬럼이 부족합니다.")

    name_col = df.columns[0]
    id_col = df.columns[1]

    def get_trimmed_mean(group):
        # 2번째 컬럼부터 끝까지 수치 데이터로 가정, 절사 평균 계산
        return group.iloc[:, 2:].apply(lambda col: round(col.sort_values().iloc[1:-1].mean(), 2))

    # 그룹화 및 계산
    df_result = df.groupby([name_col, id_col]).apply(get_trimmed_mean).reset_index()
    return df_result


def df_to_excel(df: pd.DataFrame) -> io.BytesIO:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    output.seek(0)
    return output


def get_merged_report_df(file_my, file_others) -> pd.DataFrame:
    # 1. 데이터 읽기
    df_my = pd.read_excel(file_my)
    df_others = pd.read_excel(file_others)

    name_col = df_my.columns[0]
    id_col = df_my.columns[1]

    # 2. 병합 (Inner Join)
    # suffixes를 붙여서 '항목_내가', '항목_남이'로 구분
    df_merged = pd.merge(df_my, df_others, on=[name_col, id_col], suffixes=('_내가', '_남이'))

    return df_merged


# [기능 2] 차트 생성 및 리포트 로직
def create_radar_chart_img(categories, my_view, others_view, name, student_id):
    """Matplotlib 차트를 그려서 임시 파일 경로를 반환"""

    # ★ 폰트 설정 로드
    font_prop = load_custom_font()

    # 마이너스 기호 깨짐 방지
    plt.rcParams['axes.unicode_minus'] = False

    num_vars = len(categories)
    angles = [n / float(num_vars) * 2 * pi for n in range(num_vars)]
    angles += angles[:1]

    my_view_plot = my_view + my_view[:1]
    others_view_plot = others_view + others_view[:1]

    # 객체 지향 방식 사용
    fig, ax = plt.subplots(figsize=(6, 6), subplot_kw=dict(polar=True))

    ax.set_theta_offset(pi / 2)
    ax.set_theta_direction(-1)

    # X축 (카테고리) - ★ fontproperties 적용
    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(categories, size=10, fontproperties=font_prop)

    # Y축 (점수)
    ax.set_rlabel_position(0)
    ax.set_yticks([1, 2, 3, 4, 5])
    ax.set_yticklabels(["1", "2", "3", "4", "5"], color="grey", size=7)
    ax.set_ylim(0, 5.5)

    # 데이터 플롯
    # 범례(label)에도 한글이 들어가므로 나중에 legend 설정 시 폰트 적용 필요
    ax.plot(angles, my_view_plot, linewidth=2, linestyle='solid', color='blue', label='내가 보는 나')
    ax.fill(angles, my_view_plot, color='blue', alpha=0.25)

    ax.plot(angles, others_view_plot, linewidth=2, linestyle='solid', color='red', label='남이 보는 나')
    ax.fill(angles, others_view_plot, color='red', alpha=0.1)

    # 타이틀 - ★ fontproperties 적용
    ax.set_title(f"{name} ({student_id}) 역량 분석", size=15, color='black', y=1.1, fontproperties=font_prop)

    # 범례 - ★ prop 적용
    ax.legend(loc='upper right', bbox_to_anchor=(1.3, 0.1), prop=font_prop)

    # 값 텍스트 표시 (숫자라 폰트 설정 덜 중요하지만 통일성 위해)
    for i, val in enumerate(my_view_plot[:-1]):
        ax.text(angles[i], val + 0.3, str(val), color='blue', ha='center', size=9, weight='bold')
    for i, val in enumerate(others_view_plot[:-1]):
        ax.text(angles[i], val + 0.3, str(val), color='red', ha='center', size=9, weight='bold')

    # 임시 파일로 저장
    fd, path = tempfile.mkstemp(suffix=".png")
    plt.savefig(path, bbox_inches='tight', dpi=100)
    plt.close(fig)
    os.close(fd)

    return path


def generate_report_logic(file_my, file_others) -> io.BytesIO:
    # 1. 데이터 읽기
    df_my = pd.read_excel(file_my)
    df_others = pd.read_excel(file_others)

    name_col = df_my.columns[0]
    id_col = df_my.columns[1]

    # 2. 병합
    df_merged = pd.merge(df_my, df_others, on=[name_col, id_col], suffixes=('_내가', '_남이'))
    raw_categories = df_my.columns[2:].tolist()

    # 3. 엑셀 워크북 생성
    wb = Workbook()
    wb.remove(wb.active)

    temp_images = []

    try:
        for index, row in df_merged.iterrows():
            name = row[name_col]
            student_id = str(row[id_col])

            # 데이터 추출
            my_vals = [row[f"{cat}_내가"] for cat in raw_categories]
            others_vals = [row[f"{cat}_남이"] for cat in raw_categories]

            # 차트 생성 (폰트 적용된 함수 호출)
            img_path = create_radar_chart_img(raw_categories, my_vals, others_vals, name, student_id)
            temp_images.append(img_path)

            # 시트 생성
            sheet_name = sanitize_sheet_title(student_id)
            ws = wb.create_sheet(title=sheet_name)

            # 이미지 삽입
            img = ExcelImage(img_path)
            ws.add_image(img, 'A1')

            # 표 데이터 추가
            start_row = 30
            ws.cell(row=start_row, column=1, value="역량 항목")
            ws.cell(row=start_row, column=2, value="내가 보는 점수")
            ws.cell(row=start_row, column=3, value="남이 보는 점수")

            for i, cat in enumerate(raw_categories):
                ws.cell(row=start_row + 1 + i, column=1, value=cat)
                ws.cell(row=start_row + 1 + i, column=2, value=my_vals[i])
                ws.cell(row=start_row + 1 + i, column=3, value=others_vals[i])

        # 4. 결과 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    finally:
        # 사용한 임시 이미지 파일들 삭제
        for path in temp_images:
            if os.path.exists(path):
                try:
                    os.remove(path)
                except:
                    pass
